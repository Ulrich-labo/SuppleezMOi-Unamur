from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import User
from .models import Cours,Demandes,ImputationSalaire,UserApp,ArchiveDemandeSupprime,ArchiveDemandeTraite
from django.http import HttpResponse
from openpyxl import load_workbook
from django.core.mail import send_mail
from django.contrib import admin, messages
from django.contrib.auth.hashers import make_password
from django.utils.crypto import get_random_string
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.urls import reverse_lazy

class DemandeAffichage(admin.ModelAdmin):
    """
    Custom admin class for displaying and exporting Demandes
    ...
    Attributes
    ----------
    list_display : tuple
        Tuple of fields to be displayed in the admin list view
    actions : list
        List of custom actions for the admin interface
    export_to_excel(request, queryset)
        Custom action to export selected Demandes to Excel
    """

    list_display = ('cours', 'professeur', 'suppleant', 'date_soumission', 'statut_demande')

    actions = ['export_to_excel']

    search_fields = ['cours__code','cours__titre','professeur__first_name','professeur__email','professeur__last_name']

    def export_to_excel(self, request, queryset):
        """
        Custom action to export selected Demandes to Excel
        """
        # Load the base Excel workbook as a template
        workbook = load_workbook(filename='static/Classeur1.xlsx')
        
        # Select the first sheet (Feuil1 in this example)
        assert 'Feuil1' in workbook.sheetnames, "Sheet 'Feuil1' not found in the workbook."

        # Select the first sheet (Feuil1 in this example)
        sheet = workbook['Feuil1']
    
        # Add the data
        for demande in queryset:
            row_data = [
                demande.statut_demande,
                demande.professeur.first_name,
                demande.professeur.last_name,
                demande.cours.code,
                demande.cours.titre,
                demande.cours.Programmes,
                demande.confirmation_Pour_cours_optionnel,
                demande.suppleant,
                demande.Motif,
                demande.imputation.imputation,
                demande.cpo,
                demande.Accord_suppleant,
                demande.remarque,
                demande.num_poste,
                demande.salaire,
                demande.Accord_CF,
            ]
            sheet.append(row_data)

        # Create an HTTP response for download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=demandes_excel_export.xlsx'

        # Save the workbook into the response
        workbook.save(response)

        return response

    export_to_excel.short_description = "Télécharger comme fichier Excel"



 
class CoursAffichage(admin.ModelAdmin):
    """
    Custom admin class for displaying Cours
    ...
    Attributes
    ----------
    list_display : tuple
        Tuple of fields to be displayed in the admin list view
    """
    list_display = ('code', 'titre', 'Programmes')
    search_fields = ['code', 'titre']

class ProfAffichage(admin.ModelAdmin):
    list_display = ('first_name','last_name', 'email')
    actions = [ 'send_temp_password']
    search_fields = ['last_name', 'first_name','email']
    def send_temp_password(self, request, queryset):
        """
        Custom action to send a temporary password to selected Professeurs.
        """
        for professeur in queryset:
            # generate a temporay securised passord
            temp_password = get_random_string(length=8)
            # Hash the password before to save it in the data base.
            professeur.password = make_password(temp_password)
            professeur.save(update_fields=['password'])

            # Create the token and the  uid for the reset password link 
            token = default_token_generator.make_token(professeur)
            uid = urlsafe_base64_encode(force_bytes(professeur.pk))
            password_reset_link = request.build_absolute_uri(reverse_lazy('password_reset_confirm', kwargs={'uidb64': uid, 'token': token}))
                
            
            # send the password by email.
            send_mail(
                subject='Votre mot de passe temporaire',
                message=f"Bonjour {professeur.first_name}, voici votre username:{professeur.username} et votre mot de passe temporaire : {temp_password}\nVeuillez le changer dès votre première connexion. Vous pouvez également réinitialiser directement votre mot de passe en suivant ce lien :\n{password_reset_link} Ce lien n'est valable que pour une durée limitée. Après son expiration, vous devrez vous rendre sur le site et cliquer sur 'Mot de passe oublié?' pour procéder à la réinitialisation."
,
                from_email=None,  # use the default email
                recipient_list=[professeur.email],
                fail_silently=False,
            )
        self.message_user(request, "Les mots de passe temporaires ont été envoyés avec succès.", messages.SUCCESS)

    send_temp_password.short_description = "Envoyer un mot de passe temporaire par e-mail"


# Register ArchiveDemandeTraite and ArchiveDemandeSupprime models with the custom admin class DemandeAffichage
admin.site.register(ArchiveDemandeTraite, DemandeAffichage)
admin.site.register(ArchiveDemandeSupprime, DemandeAffichage)

# Register UserApp model with the custom admin class UserAdmin
admin.site.register(UserApp,ProfAffichage)

# Register Cours model with the custom admin class CoursAffichage
admin.site.register(Cours, CoursAffichage)

# Register Demandes model with the custom admin class DemandeAffichage
admin.site.register(Demandes, DemandeAffichage)

# Register ImputationSalaire model
admin.site.register(ImputationSalaire)
